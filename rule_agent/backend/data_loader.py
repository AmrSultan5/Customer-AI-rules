"""
STEP 2 — DataLoader
Loads and normalizes all data sources.
"""

import logging
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

import pandas as pd
import yaml

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

DATA_DIR = Path(__file__).parent / "data"
RULES_FILE = DATA_DIR / "dim_rules_inventory.xlsx"
SAP_FILE   = DATA_DIR / "MDG Official Z1_AI_AGENT.xlsx"
SAP_FILE_Z11 = DATA_DIR / "MDG Official Z11.xlsx"

# Full governance repo dropped under data/
_GOVERNANCE_ROOT = DATA_DIR / "daar-governance-data-quality-processes" / "governance_data_quality_processes"
GOLDEN_DIR     = _GOVERNANCE_ROOT / "configs" / "processes" / "data_quality"  # all domains (ca / cz / it)
CUSTOM_OPS_DIR = _GOVERNANCE_ROOT / "custom_operations"

# Sheets whose first column is a row-label (Field Label / SAP Table etc.) — i.e. transposed layout
_SKIP_SHEETS = {"Z11_Transposed", "Change log", "Template Navigation", "Hier levels"}
# Row-label strings that identify the SAP Table and SAP Field rows in template sheets
_LABEL_SAP_TABLE = {"sap table", "sap\ntable"}
_LABEL_SAP_FIELD = {"sap field", "sap\nfield"}
_LABEL_FIELD_LABEL = {"field label"}
_LABEL_MDG_VAL = {"mdg validation rules", "mdg validations"}
_LABEL_REQUIRED = {"required / optional", "required/optional"}
_LABEL_FORMAT = {"sap field format"}
_LABEL_REF_TABLE = {"reference table"}

# Matches rule IDs embedded as expression literals: expression: "'RCCOMP_12.1'"
_YAML_RULE_EXPR_RE = re.compile(
    r"expression:\s+[\"']?'(RC[A-Z]+_\d+(?:[._]\d+)?)'[\"']?",
    re.IGNORECASE,
)

# Matches class XxxOperation(... BaseOperation ...): then optional docstring
_CLASS_DOC_RE = re.compile(
    r'class\s+(\w+Operation)\s*\([^)]*\):\s*(?:"""(.*?)""")?',
    re.DOTALL,
)

# ---------- helpers ----------------------------------------------------------

def _to_snake(name: str) -> str:
    """Convert any column name to lowercase snake_case."""
    s = str(name).strip()
    s = re.sub(r"[\s\-/\\.()\[\]]+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s.lower())
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {col: _to_snake(col) for col in df.columns}
    for orig, norm in mapping.items():
        if orig != norm:
            log.info("[INFO] Normalized '%s' → '%s'", orig, norm)
    return df.rename(columns=mapping)


def _find_best_column(df: pd.DataFrame, candidates: list[str], label: str) -> str:
    """Return the first column name that exists (after normalization)."""
    cols = set(df.columns)
    for c in candidates:
        if c in cols:
            log.info("[INFO] Resolved '%s' → column '%s'", label, c)
            return c
    # Fuzzy: pick any column whose name contains any candidate fragment
    for c in candidates:
        for col in cols:
            if c in col or col in c:
                log.info("[INFO] Fuzzy-resolved '%s' → column '%s'", label, col)
                return col
    raise ValueError(
        f"[ERROR] Cannot find a column mapping to '{label}'. "
        f"Available columns: {sorted(cols)}"
    )


# ---------- rules loader -----------------------------------------------------

def load_rules() -> pd.DataFrame:
    log.info("[INFO] Loading rules from %s", RULES_FILE)
    df = pd.read_excel(RULES_FILE, sheet_name="dim_rules_inventory", dtype=str)
    df = _normalize_columns(df)

    log.info("[INFO] Shape after load: %s", df.shape)
    log.info("[INFO] Columns: %s", list(df.columns))

    # Resolve logical fields
    rule_id_col = _find_best_column(
        df,
        ["rule_code", "rule_id", "id", "code"],
        "rule_id",
    )
    domain_col = _find_best_column(
        df,
        ["domain"],
        "domain",
    )
    active_col = _find_best_column(
        df,
        ["is_active", "active", "is_active_flag"],
        "is_active",
    )
    logic_col = _find_best_column(
        df,
        ["technical_definition", "rule_logic", "logic", "definition", "expression"],
        "rule_logic",
    )

    # Standardise column aliases so downstream code uses stable names
    aliases = {
        rule_id_col: "rule_id",
        domain_col: "domain",
        active_col: "is_active",
        logic_col: "rule_logic",
    }
    for src, dst in aliases.items():
        if src != dst:
            df = df.rename(columns={src: dst})

    # Keep only active rules (is_active = 1) across all domains
    before = len(df)
    df["is_active"] = pd.to_numeric(df["is_active"], errors="coerce")
    df = df[df["is_active"] == 1.0]
    log.info(
        "[INFO] Rows after is_active=1 filter: %d (dropped %d)",
        len(df),
        before - len(df),
    )
    if "domain" in df.columns:
        domain_counts = df["domain"].value_counts().to_dict()
        log.info("[INFO] Rules by domain: %s", domain_counts)

    df = df.reset_index(drop=True)
    return df


# ---------- SAP mapper loader ------------------------------------------------

def _parse_template_sheet(ws) -> list[dict]:
    """
    Parse a transposed MDG template sheet into a list of field-dicts.
    Rows 1-10 contain row-labels in column 0 and field values in columns 1+.
    Identifies each metadata row by its row-label text.
    """
    rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 12:
            break
        rows.append(list(row))

    if not rows:
        return []

    # Index each metadata row by label
    idx: dict[str, int] = {}
    for i, row in enumerate(rows):
        label = str(row[0] or "").strip().lower().replace("\n", " ").replace("  ", " ")
        if label in _LABEL_FIELD_LABEL:
            idx["label"] = i
        elif label in _LABEL_MDG_VAL:
            idx["validation"] = i
        elif label in _LABEL_REQUIRED:
            idx["required"] = i
        elif label in _LABEL_FORMAT:
            idx["format"] = i
        elif label in _LABEL_REF_TABLE:
            idx["ref_table"] = i
        elif label in _LABEL_SAP_TABLE:
            idx["sap_table"] = i
        elif label in _LABEL_SAP_FIELD:
            idx["sap_field"] = i
        elif "cchbc" in label or "field usage" in label:
            idx["desc"] = i

    # Need at minimum a SAP Table and SAP Field row to be useful
    if "sap_table" not in idx or "sap_field" not in idx:
        return []

    n_cols = max(len(r) for r in rows)
    results: list[dict] = []
    for col in range(1, n_cols):
        def _val(key: str) -> str:
            ri = idx.get(key)
            if ri is None:
                return ""
            v = rows[ri][col] if col < len(rows[ri]) else None
            s = str(v or "").strip()
            return "" if s.lower() in ("nan", "none", "") else s

        sap_table = _val("sap_table").upper()
        sap_field = _val("sap_field").upper()
        if not sap_field:
            continue

        results.append({
            "field_label": _val("label"),
            "cchbc_field_usage_comments": _val("desc")[:200],
            "mdg_validation_rules": _val("validation")[:300],
            "required_optional": _val("required"),
            "sap_field_format": _val("format"),
            "reference_table": _val("ref_table"),
            "sap_table": sap_table,
            "sap_field": sap_field,
        })

    return results


def load_sap_map() -> pd.DataFrame:
    """
    Load all SAP field metadata from both MDG Excel files.
    Primary source: Z11_Transposed (clean reference sheet).
    Additional fields: all template sheets from both files (parsed by row-label).
    Deduplicates by TABLE-FIELD, preferring Z11_Transposed entries.
    """
    import openpyxl

    all_records: list[dict] = []
    seen_keys: set[str] = set()

    # 1. Primary source: Z11_Transposed from new AI_AGENT file
    log.info("[INFO] Loading primary SAP map from %s / Z11_Transposed", SAP_FILE)
    df_primary = pd.read_excel(SAP_FILE, sheet_name="Z11_Transposed", dtype=str)
    df_primary = _normalize_columns(df_primary)
    for _, row in df_primary.iterrows():
        table = str(row.get("sap_table", "") or "").strip().upper()
        field = str(row.get("sap_field", "") or "").strip().upper()
        if not field or field in ("NAN", "NONE"):
            continue
        key = f"{table}-{field}" if table else field
        if key not in seen_keys:
            seen_keys.add(key)
            all_records.append({
                "field_label": str(row.get("field_label", "") or "").strip(),
                "cchbc_field_usage_comments": str(row.get("cchbc_field_usage_comments", "") or "").strip()[:200],
                "mdg_validation_rules": str(row.get("mdg_validation_rules", "") or "").strip()[:300],
                "required_optional": str(row.get("required_optional", "") or "").strip(),
                "sap_field_format": str(row.get("sap_field_format", "") or "").strip(),
                "reference_table": str(row.get("reference_table", "") or "").strip(),
                "sap_table": table,
                "sap_field": field,
            })
    log.info("[INFO] Z11_Transposed contributed %d primary entries", len(all_records))

    # 2. Parse all template sheets (primary AI_AGENT file + original Z11 file)
    for excel_path in [SAP_FILE, SAP_FILE_Z11]:
        if not excel_path.exists():
            log.warning("[WARNING] File not found, skipping: %s", excel_path)
            continue
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as e:
            log.warning("[WARNING] Could not open %s: %s", excel_path, e)
            continue

        for sname in wb.sheetnames:
            if sname in _SKIP_SHEETS:
                continue
            ws = wb[sname]
            try:
                records = _parse_template_sheet(ws)
            except Exception as e:
                log.warning("[WARNING] Skipping sheet '%s': %s", sname, e)
                continue

            added = 0
            for rec in records:
                table = rec["sap_table"]
                field = rec["sap_field"]
                key = f"{table}-{field}" if table else field
                if key not in seen_keys:
                    seen_keys.add(key)
                    all_records.append(rec)
                    added += 1

            if added:
                log.info("[INFO] Sheet '%s' in %s added %d new field entries",
                         sname, excel_path.name, added)

        wb.close()

    df = pd.DataFrame(all_records)
    log.info("[INFO] Total SAP field entries across all sheets: %d", len(df))
    return df


def load_reference_codes() -> dict[str, list[dict]]:
    """
    Load reference/code tables from both MDG Excel files.
    Returns dict keyed by reference-table name (e.g. 'T077D', 'TPAUM').
    Each value is a list of {code, description} dicts.
    Reference sheets are identified by having a 'Table:' row or short code columns.
    """
    import openpyxl

    # Sheets known to contain reference code data (not field-metadata templates)
    _REF_SHEETS = {
        "Recon.acc.", "Pmt Method Suppl", "Cash Mgmt Groups",
        "Ext. Identification Types", "Bus Appointment Types",
        "Mkt.Attr definitions", "Hier levels",
    }

    result: dict[str, list[dict]] = {}

    for excel_path in [SAP_FILE, SAP_FILE_Z11]:
        if not excel_path.exists():
            continue
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        except Exception:
            continue

        for sname in wb.sheetnames:
            if sname not in _REF_SHEETS:
                continue
            ws = wb[sname]
            rows = [list(r) for r in ws.iter_rows(values_only=True)
                    if any(c is not None for c in r)]

            # Find the header row (first row where most cells are non-empty)
            header_idx = 0
            for i, row in enumerate(rows):
                non_empty = sum(1 for c in row if c is not None)
                if non_empty >= 2 and i > 0:
                    header_idx = i
                    break

            # Determine reference-table name from a "Table:" row if present
            ref_table_name = sname
            for row in rows[:5]:
                if str(row[0] or "").strip().lower().startswith("table"):
                    val = str(row[1] or "").strip()
                    if val:
                        ref_table_name = val
                    break

            headers = [str(c or "").strip() for c in rows[header_idx]]
            codes: list[dict] = []
            for row in rows[header_idx + 1:]:
                entry = {headers[i]: str(row[i] or "").strip()
                         for i in range(min(len(headers), len(row)))
                         if headers[i] and str(row[i] or "").strip()}
                if entry:
                    codes.append(entry)

            if codes:
                if ref_table_name not in result:
                    result[ref_table_name] = codes
                    log.info("[INFO] Reference sheet '%s' → key '%s' (%d codes)",
                             sname, ref_table_name, len(codes))

        wb.close()

    log.info("[INFO] Loaded %d reference code tables", len(result))
    return result


# ---------- YAML loader ------------------------------------------------------

def _extract_fields_from_ops(operations: list[Any]) -> list[str]:
    """Pull column names from 'select' operations inside a YAML pipeline."""
    fields: set[str] = set()
    if not isinstance(operations, list):
        return []
    for op in operations:
        if not isinstance(op, dict):
            continue
        if op.get("kind") == "select":
            cols = op.get("params", {}).get("columns", {})
            if isinstance(cols, dict):
                fields.update(cols.keys())
                fields.update(cols.values())
            elif isinstance(cols, list):
                fields.update(cols)
        # Also grab object_name values (data sources / lineage)
    return [f for f in fields if f]


def _extract_sources_from_ops(operations: list[Any]) -> list[str]:
    sources: list[str] = []
    if not isinstance(operations, list):
        return sources
    for op in operations:
        if not isinstance(op, dict):
            continue
        if op.get("kind") == "read_dataio":
            src = op.get("params", {}).get("object_name")
            if src:
                sources.append(src)
    return sources


def load_yaml_rules() -> dict[str, dict]:
    """Return dict keyed by transform.name → parsed rule metadata."""
    result: dict[str, dict] = {}
    yaml_files = sorted(GOLDEN_DIR.rglob("*.yaml"))
    log.info("[INFO] Found %d YAML files in %s", len(yaml_files), GOLDEN_DIR)

    for yf in yaml_files:
        try:
            raw_text = yf.read_text(encoding="utf-8")
        except OSError as e:
            # File exists in OneDrive metadata but is not locally downloaded (cloud-only)
            log.warning("[WARNING] Skipping unreadable YAML %s: %s", yf.name, e)
            continue
        try:
            data = yaml.safe_load(raw_text)
        except yaml.YAMLError as e:
            log.warning("[WARNING] Skipping %s: %s", yf.name, e)
            continue

        if not isinstance(data, dict) or "transform" not in data:
            log.warning("[WARNING] Skipping %s: no 'transform' key", yf.name)
            continue

        transform = data["transform"]
        name: str = transform.get("name", yf.stem)
        operations: list = transform.get("operations") or []

        fields_used = _extract_fields_from_ops(operations)
        sources = _extract_sources_from_ops(operations)

        # Extract all rule IDs referenced as expression literals in this YAML
        rule_ids_found = list(dict.fromkeys(
            m.upper() for m in _YAML_RULE_EXPR_RE.findall(raw_text)
        ))

        # Fallback: extract SAP-style fields from any text in the YAML
        yaml_text = str(data)
        sap_fields_regex = re.findall(r"\b([A-Z][A-Z0-9]{1,9}-[A-Z][A-Z0-9_]{1,19})\b", yaml_text)

        result[name] = {
            "yaml_file": str(yf.relative_to(GOLDEN_DIR)),
            "name": name,
            "operations": operations,
            "fields_used": list(dict.fromkeys(fields_used)),
            "sources": sources,
            "sap_fields_in_yaml": list(dict.fromkeys(sap_fields_regex)),
            "custom_ops_used": _extract_custom_ops_from_ops(operations),
            "rule_ids_in_yaml": rule_ids_found,
            "description": f"Data pipeline: {name}",
        }

    log.info("[INFO] Loaded %d YAML transforms", len(result))
    return result


# ---------- yaml raw content -------------------------------------------------

@lru_cache(maxsize=256)
def get_yaml_raw(yaml_filename: str) -> str:
    """Return the raw text of a YAML file from the golden/ directory."""
    path = GOLDEN_DIR / yaml_filename
    if not path.exists():
        log.warning("[WARNING] YAML file not found: %s", path)
        return ""
    return path.read_text(encoding="utf-8")


def extract_rule_section_from_yaml(yaml_text: str, rule_id: str) -> str:
    """
    Extract the block of YAML lines that belong to a specific rule.
    Searches for the rule ID (e.g. RCCOMP_12.1) in comments and expressions,
    then returns that section plus surrounding context lines.
    Falls back to the first 3000 chars if no specific section is found.
    """
    if not yaml_text or not rule_id:
        return yaml_text[:3000] if yaml_text else ""

    # Normalise rule_id variants: RCCOMP_12.1 → rccomp_12.1, rccomp_12_1, etc.
    rid = rule_id.upper()
    rid_underscore = rid.replace(".", "_")   # RCCOMP_12_1
    rid_dot        = rid                      # RCCOMP_12.1
    patterns = [rid_dot, rid_underscore, rid.lower(), rid_underscore.lower()]

    lines = yaml_text.splitlines()
    # Find the first line containing the rule ID
    anchor = -1
    for i, line in enumerate(lines):
        if any(p in line for p in patterns):
            anchor = i
            break

    if anchor == -1:
        # No specific section found — return first 3000 chars
        return yaml_text[:3000]

    # Walk backwards to the nearest comment header (# SELECT / # ADD …)
    start = anchor
    for i in range(anchor, max(anchor - 60, -1), -1):
        if lines[i].startswith("#"):
            start = i
            break

    # Walk forward to the next rule's anchor comment or end of file
    end = min(anchor + 120, len(lines))
    for i in range(anchor + 1, len(lines)):
        if lines[i].startswith("#") and i > anchor + 5:
            # Check if the next comment belongs to a different rule
            next_comment = lines[i].upper()
            if any(p in next_comment for p in patterns):
                continue  # Still same rule, keep going
            end = i
            break

    section = "\n".join(lines[start:end])
    log.info(
        "[INFO] extract_rule_section_from_yaml: found section for %s (lines %d–%d)",
        rule_id, start, end,
    )
    return section


def find_yaml_for_rule(rule_id: str) -> dict | None:
    """
    Return the yaml metadata dict that contains rule_id in its pipeline.
    Priority:
      1. Content match — rule_id appears as an expression literal inside the YAML
      2. Name match — transform name contains the rule_id fragments (fallback)
    """
    yamls = get_yaml_rules()
    rid_upper = rule_id.upper()
    # Normalise dots/underscores for loose comparison: RCCOMP_12.1 == RCCOMP_12_1
    rid_norm = rid_upper.replace(".", "_")

    # 1. Content match: prefer the YAML that explicitly evaluates this rule
    for data in yamls.values():
        for eid in data.get("rule_ids_in_yaml", []):
            if eid.upper() == rid_upper or eid.upper().replace(".", "_") == rid_norm:
                return data

    # 2. Name heuristic fallback
    rid_lower = rule_id.lower().replace("_", "").replace(".", "").replace("-", "")
    for name, data in yamls.items():
        name_norm = name.lower().replace("_", "").replace(".", "").replace("-", "")
        if rid_lower in name_norm or name_norm in rid_lower:
            return data

    parts = [p for p in re.split(r"[_.\-]", rule_id.lower()) if len(p) > 3]
    for name, data in yamls.items():
        if any(p in name.lower() for p in parts):
            return data

    return None


# ---------- custom operations loader -----------------------------------------

def _module_key_from_path(py_file: Path) -> str:
    """Convert a .py path relative to CUSTOM_OPS_DIR to a dotted module key.
    e.g. city_standarization/geocoords_address_conformity.py
         → city_standarization.geocoords_address_conformity
    """
    rel = py_file.relative_to(CUSTOM_OPS_DIR).with_suffix("")
    return str(rel).replace("\\", "/").replace("/", ".")


def load_custom_operations() -> dict[str, dict]:
    """Scan custom_operations/ for Operation classes and extract their docstrings.
    Returns dict keyed by dotted module path fragment, e.g.:
      "city_standarization.geocoords_address_conformity" → {class_name, docstring, file}
    """
    result: dict[str, dict] = {}
    py_files = [
        f for f in CUSTOM_OPS_DIR.rglob("*.py")
        if f.name != "__init__.py"
    ]
    log.info("[INFO] Found %d custom operation source files", len(py_files))

    for py_file in py_files:
        try:
            source = py_file.read_text(encoding="utf-8")
        except Exception as e:
            log.warning("[WARNING] Could not read %s: %s", py_file, e)
            continue

        key = _module_key_from_path(py_file)
        matches = _CLASS_DOC_RE.findall(source)

        # Store the first Operation class found (usually there's only one)
        if matches:
            class_name, raw_doc = matches[0]
            docstring = " ".join(raw_doc.split()).strip() if raw_doc else ""
        else:
            # No Operation class — still index by key so YAML references resolve
            class_name = py_file.stem
            docstring = ""

        result[key] = {
            "class_name": class_name,
            "docstring": docstring,
            "file": str(py_file.relative_to(DATA_DIR)),
        }

    log.info("[INFO] Indexed %d custom operation modules", len(result))
    return result


def _extract_custom_ops_from_ops(operations: list[Any]) -> list[str]:
    """Return module-key strings for any custom operation kind found in a YAML pipeline."""
    custom_ops: list[str] = []
    if not isinstance(operations, list):
        return custom_ops
    prefix = "governance_data_quality_processes.custom_operations."
    for op in operations:
        if not isinstance(op, dict):
            continue
        kind = op.get("kind", "")
        if not isinstance(kind, str) or prefix not in kind:
            continue
        # Strip the prefix and trailing .ClassName to get the module key
        fragment = kind[kind.index(prefix) + len(prefix):]
        parts = fragment.rsplit(".", 1)
        module_key = parts[0] if len(parts) == 2 else fragment
        if module_key and module_key not in custom_ops:
            custom_ops.append(module_key)
    return custom_ops


# ---------- cached entry points ----------------------------------------------

@lru_cache(maxsize=1)
def get_rules() -> pd.DataFrame:
    return load_rules()


@lru_cache(maxsize=1)
def get_sap_map() -> pd.DataFrame:
    return load_sap_map()


@lru_cache(maxsize=1)
def get_reference_codes() -> dict[str, list[dict]]:
    return load_reference_codes()


@lru_cache(maxsize=1)
def get_yaml_rules() -> dict[str, dict]:
    return load_yaml_rules()


@lru_cache(maxsize=1)
def get_custom_operations() -> dict[str, dict]:
    return load_custom_operations()


@lru_cache(maxsize=128)
def get_custom_op_source(module_key: str, max_chars: int = 6000) -> str:
    """Return the source code of a custom operation module, truncated to max_chars.

    The path is resolved strictly via the get_custom_operations() index — never
    from raw user/LLM input — so there is no path-traversal risk.
    Returns "" for unknown keys or read failures.
    """
    meta = get_custom_operations().get(module_key)
    if not meta:
        log.warning("[WARNING] get_custom_op_source: unknown module key %r", module_key)
        return ""
    path = DATA_DIR / meta["file"]
    try:
        source = path.read_text(encoding="utf-8")
    except Exception as e:
        log.warning("[WARNING] get_custom_op_source: could not read %s: %s", path, e)
        return ""
    if len(source) > max_chars:
        source = source[:max_chars] + "\n# … truncated"
    return source


def clear_caches() -> None:
    """Drop all cached data so the next access re-reads from disk."""
    for fn in (
        get_rules, get_sap_map, get_reference_codes,
        get_yaml_rules, get_custom_operations, get_custom_op_source,
    ):
        fn.cache_clear()


def reload_all(descriptor: Any = None) -> dict:
    """Reload all data sources from disk, validating BEFORE swapping the caches.

    Loads fresh copies first and runs schema validation on them; only when the
    new data is good are the caches cleared and re-warmed. A broken Excel/YAML
    edit therefore raises here while the app keeps serving the old data.

    `descriptor` is optional (Phase 2): when passed (a KBDescriptor), validation
    runs against its field_map-declared required columns via
    schema_validator.validate_against_descriptor. Existing callers that pass
    nothing (main.py's /admin/reload) keep the original hardcoded-column
    validation — same effective required set for customer_sap either way.
    """
    from schema_validator import validate_against_descriptor, validate_rules, validate_sap

    new_rules = load_rules()
    new_sap = load_sap_map()
    if descriptor is not None:
        validate_against_descriptor(new_rules, new_sap, descriptor)
    else:
        validate_rules(new_rules)
        validate_sap(new_sap)
    new_yamls = load_yaml_rules()
    new_ops = load_custom_operations()

    clear_caches()
    get_rules()
    get_sap_map()
    get_yaml_rules()
    get_custom_operations()

    log.info(
        "[INFO] Data reloaded: %d rules, %d pipelines, %d custom ops, %d SAP fields",
        len(new_rules), len(new_yamls), len(new_ops), len(new_sap),
    )
    return {
        "rules_loaded": len(new_rules),
        "yaml_pipelines": len(new_yamls),
        "custom_ops": len(new_ops),
        "sap_fields": len(new_sap),
    }


# ---------- cross-rule reference resolver ------------------------------------

_RULE_REF_RE = re.compile(r"\b(RC[A-Z]+_\d+(?:[._]\d+)?)\b", re.IGNORECASE)


def get_referenced_rules(rule_id: str) -> list[dict]:
    """Return details of rules referenced by rule_id via dependent_on or rule_logic.

    Each entry: {rule_id, rule_description, rule_logic, table_name_checked,
                 column_name_checked, quality_category, source}
    where source is 'dependent_on' or 'logic'.
    """
    rules = get_rules()
    match = rules[rules["rule_id"].str.upper() == rule_id.upper()]
    if match.empty:
        return []

    row = match.iloc[0]
    rid_upper = rule_id.upper()
    seen: set[str] = set()
    refs: list[dict] = []

    def _collect(text: str, source: str) -> None:
        for m in _RULE_REF_RE.finditer(str(text or "")):
            ref_id = m.group(1).upper()
            if ref_id == rid_upper or ref_id in seen:
                continue
            seen.add(ref_id)
            ref_match = rules[rules["rule_id"].str.upper() == ref_id]
            if ref_match.empty:
                # Reference exists in text but not in active rules
                refs.append({"rule_id": ref_id, "source": source, "active": False})
            else:
                ref_row = ref_match.iloc[0]
                refs.append({
                    "rule_id": ref_id,
                    "rule_description": str(ref_row.get("rule_description", "") or ""),
                    "rule_logic": str(ref_row.get("rule_logic", "") or ""),
                    "table_name_checked": str(ref_row.get("table_name_checked", "") or ""),
                    "column_name_checked": str(ref_row.get("column_name_checked", "") or ""),
                    "quality_category": str(ref_row.get("quality_category", "") or ""),
                    "source": source,
                    "active": True,
                })

    _collect(row.get("dependent_on", ""), "dependent_on")
    _collect(row.get("rule_logic", ""), "logic")
    return refs


# ---------- sanity check -----------------------------------------------------

if __name__ == "__main__":
    rules = get_rules()
    print(f"\nRules shape: {rules.shape}")
    print(f"Columns: {list(rules.columns)}")
    print(f"\nSample rule_id values: {rules['rule_id'].head(5).tolist()}")

    sap = get_sap_map()
    print(f"\nSAP map shape: {sap.shape}")
    print(f"SAP columns: {list(sap.columns)}")

    yamls = get_yaml_rules()
    print(f"\nYAML rules loaded: {len(yamls)}")
    first_key = next(iter(yamls))
    print(f"First YAML key: {first_key}")
    print(f"First YAML fields_used (first 5): {yamls[first_key]['fields_used'][:5]}")
